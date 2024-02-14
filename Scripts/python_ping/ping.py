import ctypes
import subprocess
import os
import sys
import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, Color

def is_admin():
    try:
        import ctypes
        return ctypes.windll.shell32.IsUserAnAdmin() != 0
    except AttributeError:
        return os.getuid() == 0

def ping_ip(ip_address):
    try:
        output = subprocess.check_output(["ping", "-n", "1", ip_address], stderr=subprocess.STDOUT, universal_newlines=True)
        print(output)  # Muestra la salida del comando ping
        for line in output.split("\n"):
            if "tiempo=" in line:
                time_ms = line.split("tiempo=")[1].split("ms")[0].strip()
                return True, time_ms
        return False, "0"
    except subprocess.CalledProcessError as e:
        print(e.output)  # Muestra la salida en caso de error
        return False, "0"

def main():
    if not is_admin():
        print("Este script requiere privilegios de administrador.")
        sys.exit(1)

    if not os.path.exists("ips.txt"):
        print("El archivo ips.txt no existe.")
        sys.exit(1)
    
    wb = Workbook()
    ws = wb.active
    ws.append(["IP", "Estado", "Tiempo (ms)", "Fecha y Hora"])
    ws['B1'] = 'Estado'
    ws['B2'] = 'Verdadero'
    ws['B3'] = 'Falso'
    
    with open("ips.txt", "r") as file:
        ips = file.read().splitlines()

    for ip in ips:
        is_up, time_ms = ping_ip(ip)
        current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if is_up:
            print(f"{ip} está arriba, tiempo de respuesta: {time_ms}ms")
            ws.append([ip, "Verdadero", time_ms, current_time])
            ws.conditional_formatting.add('B1:B2',
            CellIsRule(operator='equal', formula=['"Verdadero"'], fill=PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')))
           
        else:
            print(f"{ip} no responde.")
            ws.append([ip, "Falso", time_ms, current_time])
            ws.conditional_formatting.add('B1:B2',
            CellIsRule(operator='equal', formula=['"Falso"'], fill=PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')))
            
    
    
    
    
    wb.save("resultados_ping.xlsx")
    print("Resultados exportados a resultados_ping.xlsx.")

if __name__ == "__main__":
    main()











































































